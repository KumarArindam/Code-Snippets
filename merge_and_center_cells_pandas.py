import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment

def merge_and_center_by_column(dataframe, column_to_merge):
    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active

    # Convert the Pandas DataFrame data to the Excel sheet
    for r_idx, row_data in enumerate(dataframe.iterrows(), start=1):
        _, data = row_data
        for c_idx, value in enumerate(data, start=1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)

    # Define a dictionary to track merged cells
    merged_cells = {}

    # Iterate through the DataFrame to find and merge cells with the same value in the specified column
    for row in range(2, len(dataframe) + 2):
        cell_value = worksheet.cell(row=row, column=dataframe.columns.get_loc(column_to_merge) + 1).value
        if cell_value not in merged_cells:
            merged_cells[cell_value] = [row]
        else:
            merged_cells[cell_value].append(row)

    # Merge and center the cells
    for cell_value, row_indices in merged_cells.items():
        if len(row_indices) > 1:
            for col_idx in range(1, len(dataframe.columns) + 1):
                start_row = row_indices[0]
                end_row = row_indices[-1]
                worksheet.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
                worksheet.cell(row=start_row, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')

    # Save the Excel file
    workbook.save('merged_and_centered.xlsx')

# Example usage:
data = {
    'Category': ['A', 'A', 'B', 'B', 'B'],
    'Value1': [10, 20, 15, 30, 25],
    'Value2': [50, 60, 70, 80, 90]
}

df = pd.DataFrame(data)

merge_and_center_by_column(df, 'Category')
